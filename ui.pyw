#coding: utf8

import wx, sys
from openpyxl import load_workbook

class MyFrame(wx.Frame):
    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, title = title, size = (520, 420))
        panel = wx.Panel(self)

        self.sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.sizer.Add(panel, 1, wx.EXPAND)

        btnLocFreecad = wx.Button(panel, wx.ID_ANY, '请选择FreeCAD.exe所在目录')
        btnLocFcstd = wx.Button(panel, wx.ID_ANY, '请选择FCStd文件')
        btnLocXlsx = wx.Button(panel, wx.ID_ANY, '请选择Excel文件')
        self.lstSheetname = wx.Choice(panel)
        btnLocOutput = wx.Button(panel, wx.ID_ANY, '请选择输出目录')
        self.locFreecad = wx.TextCtrl(panel)
        self.locFcstd = wx.TextCtrl(panel)
        self.locXlsx = wx.TextCtrl(panel)
        self.locOutput = wx.TextCtrl(panel)
        self.locOutput.SetLabel('output')
        self.locFreecad.SetLabel('C:\\Program Files\\FreeCAD 0.17\\bin')
        self.btnOK = wx.Button(panel, wx.ID_ANY, "开始转换")
        self.btnOK.Enable(False)

        sz1 = wx.BoxSizer(wx.HORIZONTAL)
        sz1.Add(btnLocFreecad, 0, wx.EXPAND)
        sz1.Add(self.locFreecad, 1, wx.EXPAND)
    
        sz2 = wx.BoxSizer(wx.HORIZONTAL)
        sz2.Add(btnLocFcstd, 0, wx.EXPAND)
        sz2.Add(self.locFcstd, 1, wx.EXPAND)

        sz3 = wx.BoxSizer(wx.HORIZONTAL)
        sz3.Add(btnLocXlsx, 0, wx.EXPAND)
        sz3.Add(self.locXlsx, 1, wx.EXPAND)
        sz3.Add(self.lstSheetname, 0, wx.EXPAND)

        sz4 = wx.BoxSizer(wx.HORIZONTAL)
        sz4.Add(btnLocOutput, 0, wx.EXPAND)
        sz4.Add(self.locOutput, 1, wx.EXPAND)
        sz4.Add(self.btnOK)

        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(sz1, 0, wx.EXPAND)
        sizer.AddSpacer(5)
        sizer.Add(sz2, 0, wx.EXPAND)
        sizer.AddSpacer(5)
        sizer.Add(sz3, 0, wx.EXPAND)
        sizer.AddSpacer(5)
        sizer.Add(sz4, 0, wx.EXPAND)

        panel.SetSizer(sizer)
        panel.SetAutoLayout(1)
        sizer.Fit(panel)

        self.Bind(wx.EVT_BUTTON, self.OnFreecad, btnLocFreecad)
        self.Bind(wx.EVT_BUTTON, self.OnFcstd, btnLocFcstd)
        self.Bind(wx.EVT_BUTTON, self.OnXlsx, btnLocXlsx)
        self.Bind(wx.EVT_BUTTON, self.OnOutput, btnLocOutput)
        self.Bind(wx.EVT_BUTTON, self.OnOK, self.btnOK)

        self.Show(True)

    def OnOK(self, event):
        fcstdPath = self.locFcstd.GetLabel()
        xlsxPath = self.locXlsx.GetLabel()

        import main
        main.Path_Output = self.locOutput.GetLabel()
        main.Path_FreeCAD = self.locFreecad.GetLabel()
        sheetname = self.lstSheetname.GetString(self.lstSheetname.GetSelection())
        main.main(fcstdPath, xlsxPath, sheetname)
        dlg = wx.MessageDialog(self, "转换完成！", "转换结果", wx.OK)
        dlg.ShowModal()

    def Check(self):
        freeCadPath = self.locFreecad.GetLabel()
        fcstdPath = self.locFcstd.GetLabel()
        xlsxPath = self.locXlsx.GetLabel()
        outputPath = self.locOutput.GetLabel()
        sheetname = self.lstSheetname.GetString(self.lstSheetname.GetSelection())
        if len(freeCadPath) == 0 or len(fcstdPath) == 0 or len(xlsxPath) == 0 or len(outputPath) == 0 or len(sheetname) == 0:
            self.btnOK.Enable(False)
        else:
            self.btnOK.Enable(True)

    def OnFreecad(self, event):
        dlg = wx.DirDialog(self, "选择FreeCAD安装目录",style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            self.locFreecad.SetLabel(dlg.GetPath())
        dlg.Destroy()
        self.Check()

    def OnFcstd(self, event):
        filesFilter = "FreeCAD工程文件 (*.FCStd)|*.FCStd"
        fileDialog = wx.FileDialog(self, message ="选择FCStd文件", wildcard = filesFilter, style = wx.FD_OPEN)
        dialogResult = fileDialog.ShowModal()
        if dialogResult !=  wx.ID_OK:
            return
        path = fileDialog.GetPath()
        self.locFcstd.SetLabel(path)
        self.Check()

    def OnXlsx(self, event):
        def GetSheetNameFromXls(fpath):
            wb = load_workbook(filename = fpath)
            return list(wb.sheetnames)

        filesFilter = "Excel表格文件 (*.xlsx)|*.xlsx"
        fileDialog = wx.FileDialog(self, message ="选择Excel文件", wildcard = filesFilter, style = wx.FD_OPEN)
        dialogResult = fileDialog.ShowModal()
        if dialogResult !=  wx.ID_OK:
            return
        path = fileDialog.GetPath()
        self.locXlsx.SetLabel(path)
        names = GetSheetNameFromXls(path)
        self.lstSheetname.SetItems(names)
        if len(names) > 0:
            self.lstSheetname.SetSelection(0)
        self.Check()

    def OnOutput(self, event):
        dlg = wx.DirDialog(self, "选择输出目录",style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            self.locOutput.SetLabel(dlg.GetPath())
        dlg.Destroy()
        self.Check()

def main():
    app = wx.App(redirect = True)
    frame = MyFrame(None, "FCStd转换工具")
    app.MainLoop()

if __name__ == '__main__':
    main()
