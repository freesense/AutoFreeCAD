#coding: utf8

###########################################################################################
# 请不要使用 记事本 来打开本文件！如果已经打开，请一定不要保存本文件！
# 请不要使用 记事本 来打开本文件！如果已经打开，请一定不要保存本文件！！
# 请不要使用 记事本 来打开本文件！如果已经打开，请一定不要保存本文件！！！
# 可以使用 ultraedit/notepad++/vim 等完美兼容utf8编码的编辑器来编辑本文件
#
# 使用前请先安装：FreeCAD, python2.7
# 安装python2.7完成后，请执行：pip install openpyxl
# 最后将本文件置于一新建目录下，例如"d:\rongweisoft"
#
# 本程序是命令行程序，所有操作均在本文件所在目录执行
# 本程序运行需要两个输入文件，输出两个文件，输出文件军保存在下面配置的输出文件目录下
# 输入文件1：fcstd模板文件
# 输入文件2：xlsx数据文件
# 输出文件1：fcstd模型文件
# 输出文件2：step模型文件
# 进入命令行窗口，输入：
# d:\
# cd rongweisoft
# main.py 输入文件1 输入文件2 输入文件2的执行sheet名称
# 最后检查输出文件是否符合要求
#
###########################################################################################

# FreeCAD.exe安装后的绝对路径
Path_FreeCAD = 'C:\\Program Files\\FreeCAD 0.17\\bin'
# 输出文件目录，可使用相对和绝对路径
Path_Output = 'output'

import os, sys
import zipfile
from openpyxl import load_workbook

def main(infile, xlsfile, sheetname):
    if not os.path.exists(Path_Output):
        os.mkdir(Path_Output)

    sys.path.append(Path_FreeCAD)
    import FreeCAD, Import

    FreeCAD.open(infile)
    App = FreeCAD
    docs = App.listDocuments()
    docName = docs.keys()[0]
    App.setActiveDocument(docName)
    doc = App.getDocument(docName)
    rootObjs = doc.RootObjects

    wb = load_workbook(filename = xlsfile)
    if sheetname not in set(wb.sheetnames):
        print 'Sheet name: %s not in file: %s.' % (sheetname, xlsfile)
        return

    def each(name, idx, value):
        try:
            obj = doc.getObject(name)
            if idx == 'Width':
                obj.Width = value
            elif idx == 'Height':
                obj.Height = value
            elif idx == 'Length':
                obj.Length = value
            else:
                obj.setDatum(idx, App.Units.Quantity(value))
        except Exception, e:
            print e

    sh = wb[sheetname]
    for name, idx, v in sh.values:
        each(name, idx, v)
    doc.recompute()

    _, fname = os.path.split(infile)
    fname, _ = os.path.splitext(fname)
    targetName = os.path.join(Path_Output, '%s_%s.FCStd' % (fname, sheetname))
    stepName = os.path.join(Path_Output, '%s_%s.step' % (fname, sheetname))
    doc.saveAs(targetName)

    def mergeGui():
        f1 = zipfile.ZipFile(infile)
        old_files = f1.namelist()
        f2 = zipfile.ZipFile(targetName)
        new_files = set(f2.namelist())
        f2.close()

        for f in old_files:
            if f not in new_files:
                f1.extract(f, 'tmp')
        f1.close()

        f = zipfile.ZipFile(targetName, 'a', zipfile.ZIP_DEFLATED)
        os.chdir('tmp')
        for _, _, files in os.walk('.'):
            for name in files:
                f.write(name)
                os.remove(name)
        f.close()
        os.chdir('..')

        os.removedirs('tmp')

    mergeGui()
    Import.export(rootObjs, stepName)

def ui_main():
    import ui
    ui.main()

def console_main():
    if len(sys.argv) < 3:
        print 'Command line: main.py fcstd_file_path xls_file_path sheet_name'
        return
    main(sys.argv[-3], sys.argv[-2], sys.argv[-1])

if __name__ == '__main__':
    console_main()
